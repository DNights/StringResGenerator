import os
from io import StringIO
from openpyxl import load_workbook

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)

def main():
    print("Start Generate String Res")

    load_wb = load_workbook("./string_generate.xlsx", data_only=True)
    load_ws = load_wb['stringsheet']

    #xlsx get string_id
    string_id_row = 6
    string_id_col = 2
    string_ids = []

    for col_data in load_ws.iter_cols(min_col=string_id_col, max_col=string_id_col, min_row=string_id_row):
        for cell in col_data:
            value = cell.value
            if(value == None):
                break
            else:
                string_ids.append(value)

    # print(string_ids)

    #xlsx get res_file_names
    res_file_names_row = 3
    res_file_names_col = 4
    res_file_names = []

    for row_data in load_ws.iter_rows(min_col=res_file_names_col, min_row=res_file_names_row, max_row=res_file_names_row):
        for cell in row_data:
            value = cell.value
            if(value == None):
                break
            else:
                res_file_names.append(value)

    # print(res_file_names)

    #parser
    start_row_index = 6
    start_col_index = 4
    string_ids_index = 0
    res_file_names_index = 0 

    for col_data in load_ws.iter_cols(min_col= start_col_index, max_col=len(res_file_names) + start_col_index -1, min_row=start_row_index):

        print("start read " + res_file_names[res_file_names_index])

        contents = "//자동으로 생성된 파일입니다. 수정하지 마세요.\n"

        for cell in col_data:
            value = cell.value

            if(value == None):
                if(string_ids_index >= len(string_ids) -1):
                    print("end read " + res_file_names[res_file_names_index])
                    break
                else:
                    continue
            else:
                contents += '\"' + string_ids[string_ids_index] + '\" = \"' + value + '\"\n'
                # print("string_ids[" + str(string_ids_index) + "] = " + string_ids[string_ids_index] + " / " + value)
            string_ids_index = string_ids_index + 1
            
        string_ids_index = 0
        
        path = './'+res_file_names[res_file_names_index]

        createFolder(path)
        print ("create folder path - " + path)

        with open(path + "/Localizable.strings", "w") as f:
            f.write(contents)

        print ("generate string xml path - " + path + "/Localizable.strings")

        res_file_names_index = res_file_names_index + 1           
            

    print("Success Generate String Res")

if __name__ == '__main__':
    main()